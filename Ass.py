from genericpath import exists
from tracemalloc import StatisticDiff
from openpyxl import Workbook
from openpyxl import load_workbook
import openpyxl
from os.path import exists
active = True
path = "datasheet.xlsx" 

dic = {
  "StudentNo":  [],  
  "Python": [],
  "SQL": [],
  "HTML":  [],
  "JavaScript": [],
  "ASP.Net": [],
  "C#": []
}

#creates a new workbook
if exists( path ) == True:
    wb = load_workbook( path )
else:
    wb = Workbook()
    wb.save(filename = path )
    wb = load_workbook( path )
#Gets the first active worksheet

sheet = wb["Sheet"]
#Renaming the sheet



row_count = sheet.max_row
column_count = sheet.max_column
# for i in range(1, row_count + 1):
 #   for j in range(1, column_count + 1):
 #       data = sheet.cell(row=i, column=j).value
 #       print(  data  ) 

for i in range(1, row_count + 1):
    StudentNo = []
    Python= []
    SQL = []
    HTML =  []
    JavaScript = []
    ASP = []
    C = []

    StudentNo.append( sheet.cell(row=i, column= 1 ).value )
    Python.append( sheet.cell(row=i, column= 2 ).value )
    SQL.append( sheet.cell(row=i, column= 3 ).value )
    HTML.append( sheet.cell(row=i, column= 4 ).value )
    JavaScript.append( sheet.cell(row=i, column= 5 ).value )
    ASP.append( sheet.cell(row=i, column= 6 ).value )
    C.append( sheet.cell(row=i, column= 7 ).value )

    dic["StudentNo"].append( StudentNo )
    dic["Python"].append( Python )
    dic["SQL"].append( SQL )
    dic["HTML"].append( HTML )
    dic["JavaScript"].append( JavaScript )
    dic["ASP.Net"].append( ASP )
    dic["C#"].append( C )

    StudentNo = []
    Python= []
    SQL = []
    HTML =  []
    JavaScript = []
    ASP = []
    C = []

Stds = [ ( "StudentNo", "Python", "SQL", "HTML", "JavaScript", "ASP.Net", "C#"  ) ]

def createarry():
    global Stds
    global sheet
    Stds = [ ( "StudentNo", "Python", "SQL", "HTML", "JavaScript", "ASP.Net", "C#"  ) ]
    if  len( dic["StudentNo"] )  != 0 :
        
        for i in range(  len( dic["StudentNo"] ) ):
            listtemp = []
            listtemp.append( dic["StudentNo"][i] )
            listtemp.append( dic["Python"][i] ) 
            listtemp.append( dic["SQL"][i])
            listtemp.append( dic["HTML"][i] )
            listtemp.append( dic["JavaScript"][i] )
            listtemp.append( dic["ASP.Net"][i] )
            listtemp.append( dic["C#"][i] )
            
            Stds.append( listtemp )
            listtemp = []
    
            
    

def writetoel():
    global sheet
    global wb
    sheet.delete_rows(1, sheet.max_row+1)
    createarry()
    for i in Stds :
        sheet.append(i)
    wb.save(filename = path)

def entermark():
    sname = input("Enter student name :")
    pym = int(input("Enter python marke :"))
    SQL= int(input("Enter sql marke :"))
    HTML= int(input("Enter html marke :"))
    JavaScript= int(input("Enter javascript marke :"))
    ASP= int(input("Enter asp.net marke :"))
    C= int(input("Enter c# marke :"))
    
    dic["StudentNo"].append( sname )
    dic["Python"].append( pym )
    dic["SQL"].append( SQL )
    dic["HTML"].append( HTML )
    dic["JavaScript"].append( JavaScript )
    dic["ASP.Net"].append( ASP )
    dic["C#"].append( C )


def displaymark():
    createarry()
    for i in Stds:
        print( i )

#def  delmark():
#    sname = input("Enter student name :")
#    dic

#def resofAstud():

#def dispResOfAllStd():


# Set up the while loop.
while active:
    resp = input("Please press \n1 to Enter Marks \n2 to Dsplay Marks \n3 to Delete Marks \n4 to Display Result of a Given Student \n5 to Display Results of all the Students \n6 to Exit ...\n")
    
    if resp == "1":
        entermark()
 
    elif resp == "2":
        displaymark()
    
    elif resp == "3":
        #delmark()
        print("not implemented")

    elif resp == "4":
        #resofAstud()
        print("not implemented")
    
    elif resp == "5":
        #dispResOfAllStd()
        print("not implemented")

    elif resp == "6":
        writetoel()
        active = False        
    
    else:
        print("\nPlease please please , enter a valid input ....\n")

        




